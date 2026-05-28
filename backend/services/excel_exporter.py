"""Excel export service using openpyxl."""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _style_header(ws, row: int, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def export_portfolio(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Портфель"

    headers = ["ID Сделки", "Клиент", "Тип", "Статус", "Сумма", "Дата создания"]
    _style_header(ws, 1, headers)

    for i, row in enumerate(rows, 2):
        ws.cell(i, 1, str(row.get("id", "")))
        ws.cell(i, 2, row.get("client_name", ""))
        ws.cell(i, 3, row.get("type", ""))
        ws.cell(i, 4, row.get("status", ""))
        ws.cell(i, 5, float(row.get("total", 0)))
        ws.cell(i, 6, str(row.get("created_at", "")))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_overdue(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Просрочки"

    headers = [
        "ID Дела", "Клиент", "Телефон", "Сумма долга",
        "Дней просрочки", "Статус дела", "Сотрудник СБ",
    ]
    _style_header(ws, 1, headers)

    for i, row in enumerate(rows, 2):
        ws.cell(i, 1, str(row.get("case_id", "")))
        ws.cell(i, 2, row.get("client_name", ""))
        ws.cell(i, 3, row.get("phone", ""))
        ws.cell(i, 4, float(row.get("total_debt", 0)))
        ws.cell(i, 5, row.get("days_overdue", 0))
        ws.cell(i, 6, row.get("status", ""))
        ws.cell(i, 7, row.get("sb_user_name", "Не назначен"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_audit_log(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Log"

    headers = ["ID", "Пользователь", "Действие", "Сущность", "ID сущности", "IP", "Дата/Время"]
    _style_header(ws, 1, headers)

    for i, row in enumerate(rows, 2):
        ws.cell(i, 1, row.get("id", ""))
        ws.cell(i, 2, str(row.get("user_id", "")))
        ws.cell(i, 3, row.get("action", ""))
        ws.cell(i, 4, row.get("entity", ""))
        ws.cell(i, 5, str(row.get("entity_id", "")))
        ws.cell(i, 6, row.get("ip", ""))
        ws.cell(i, 7, str(row.get("created_at", "")))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_payment_schedule(deal: dict, schedules: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "График платежей"

    ws.cell(1, 1, f"Сделка: {deal.get('id')}")
    ws.cell(2, 1, f"Клиент: {deal.get('client_name')}")
    ws.cell(3, 1, f"Тип: {deal.get('type')}")
    ws.cell(4, 1, f"Итого: {deal.get('total')}")

    headers = ["№", "Дата платежа", "Сумма", "Оплачено", "Остаток", "Статус"]
    _style_header(ws, 6, headers)

    for i, row in enumerate(schedules, 7):
        amount = float(row.get("amount", 0))
        paid = float(row.get("paid_amount", 0))
        ws.cell(i, 1, row.get("installment_number", ""))
        ws.cell(i, 2, str(row.get("due_date", "")))
        ws.cell(i, 3, amount)
        ws.cell(i, 4, paid)
        ws.cell(i, 5, round(amount - paid, 2))
        ws.cell(i, 6, row.get("status", ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
